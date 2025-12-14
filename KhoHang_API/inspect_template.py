"""
inspect_template.py - Kiểm tra cấu trúc template Excel
"""
import sys
sys.path.insert(0, '.')

from openpyxl import load_workbook

def inspect_template(filepath):
    print(f"\n{'='*60}")
    print(f"Inspecting: {filepath}")
    print(f"{'='*60}\n")
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    print(f"Sheet name: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")
    
    # Print merged cells
    print(f"\nMerged cells:")
    for mcr in ws.merged_cells.ranges:
        print(f"  {mcr}")
    
    # Print first 20 rows
    print("\n\nFirst 50 rows content:")
    for row in range(1, 51):
        row_data = []
        for col in 'ABCDEFGH':
            cell = ws[f'{col}{row}']
            val = cell.value if hasattr(cell, 'value') else f"<merged:{type(cell).__name__}>"
            if val is not None:
                row_data.append(f"{col}{row}:{val}")
        if row_data:
            print(f"  Row {row:2}: {row_data}")

# Check both templates
inspect_template('templates/Mau_Phieu_Xuat_TEMPLATE.xlsx')
