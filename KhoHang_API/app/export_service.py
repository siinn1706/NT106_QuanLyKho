"""
export_service.py - Service xuất phiếu nhập/xuất kho ra Excel và PDF

Flow:
1. Nhận VoucherExportData (JSON)
2. Load template Excel (.xlsx)
3. Fill data vào template
4. Trả về file Excel hoặc convert sang PDF
"""

import os
import io
from datetime import datetime
from typing import Optional, List
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from pydantic import BaseModel
from .number_to_words_vi import amount_to_vietnamese_words


# ============================================
# DATA MODELS (Schemas)
# ============================================

class VoucherItem(BaseModel):
    """Chi tiết một dòng hàng hóa"""
    sku: str                    # Mã hàng
    name: str                   # Tên hàng
    unit: str                   # Đơn vị tính
    qty_doc: float              # Số lượng chứng từ
    qty_actual: float           # Số lượng thực tế
    unit_price: float           # Đơn giá


class VoucherExportRequest(BaseModel):
    """
    Request body để xuất phiếu
    
    JSON Contract:
    {
        "voucher_type": "PN" | "PX",
        "voucher_no": "PN-000123",
        "voucher_date": "2025-12-13",
        "partner_name": "Nhà cung cấp A",
        "invoice_no": "HD001",
        "invoice_date": "2025-12-13",
        "warehouse_code": "K01",
        "warehouse_location": "Kho chính",
        "attachments": "1 phiếu",
        "tax_rate": 8,
        "prepared_by": "Nguyễn Văn A",
        "receiver": "Trần Văn B",
        "storekeeper": "Lê Thị C",
        "director": "Phạm Văn D",
        "items": [
            {
                "sku": "SP-001",
                "name": "Sản phẩm A",
                "unit": "Cái",
                "qty_doc": 10,
                "qty_actual": 10,
                "unit_price": 100000
            }
        ]
    }
    """
    voucher_type: str           # "PN" (Phiếu Nhập) hoặc "PX" (Phiếu Xuất)
    voucher_no: str             # Số phiếu
    voucher_date: str           # Ngày phiếu (ISO format hoặc dd/MM/yyyy)
    partner_name: str           # Tên đối tác (NCC/Khách hàng)
    invoice_no: Optional[str] = None
    invoice_date: Optional[str] = None
    warehouse_code: str
    warehouse_location: Optional[str] = None
    attachments: Optional[str] = None
    tax_rate: Optional[float] = 0  # Thuế suất (%): 0, 5, 8, 10...
    prepared_by: Optional[str] = None
    receiver: Optional[str] = None
    storekeeper: Optional[str] = None
    director: Optional[str] = None
    items: List[VoucherItem]


# ============================================
# HELPER FUNCTIONS
# ============================================

def _format_date(date_str: str) -> str:
    """Format date string to dd/MM/yyyy"""
    try:
        # Try ISO format
        date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return date.strftime('%d/%m/%Y')
    except:
        # Return as-is if already formatted
        return date_str


def _format_number(num: float) -> str:
    """Format number with thousand separator"""
    return f"{num:,.0f}".replace(",", ".")


def _get_template_path(voucher_type: str) -> str:
    """Get template file path based on voucher type"""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_dir = os.path.join(base_dir, 'templates')
    
    if voucher_type == 'PN':
        return os.path.join(template_dir, 'Mau_Phieu_Nhap_TEMPLATE.xlsx')
    else:
        return os.path.join(template_dir, 'Mau_Phieu_Xuat_TEMPLATE.xlsx')


def _generate_filename(data: VoucherExportRequest, extension: str) -> str:
    """
    Generate filename: <voucher_type>_<warehouse_code>_<YYYYMM>_<voucher_no>.<ext>
    VD: PN_K01_202512_PN-000123.xlsx
    """
    try:
        date = datetime.fromisoformat(data.voucher_date.replace('Z', '+00:00'))
    except:
        date = datetime.now()
    
    yyyy = date.year
    mm = str(date.month).zfill(2)
    
    # Sanitize voucher_no for filename
    voucher_no_safe = data.voucher_no.replace('/', '-').replace('\\', '-')
    warehouse_safe = data.warehouse_code.replace('/', '-').replace('\\', '-')
    
    return f"{data.voucher_type}_{warehouse_safe}_{yyyy}{mm}_{voucher_no_safe}.{extension}"


# ============================================
# EXCEL EXPORT
# ============================================

def export_to_excel(data: VoucherExportRequest) -> tuple[io.BytesIO, str]:
    """
    Export voucher data to Excel using template
    
    Args:
        data: VoucherExportRequest with all voucher data
    
    Returns:
        Tuple of (BytesIO buffer, filename)
    
    Template Contract (Mau_Phieu_Nhap_TEMPLATE.xlsx):
    - Header cells:
        - F7 (merged D7:E7 label "Số phiếu:"): Voucher number
        - D2 (merged D2:F2): Date (has =TODAY() formula, we override)
        - C9: Partner name (Họ và tên người nhận)
        - E10: Invoice number (Theo hóa đơn số)
        - C11: Warehouse (Nhận tại kho)
        - F11: Warehouse location (Địa điểm)
    - Items table: rows 16-33 (18 rows max)
        - A: STT
        - B: Tên hàng (merged B:C)
        - D: Mã số
        - E: Đơn vị tính
        - F: SL Chứng từ
        - G: SL Thực nhập
        - H: Đơn giá
        - I: Thành tiền (formula)
    - Summary:
        - Row 36: Tổng số tiền bằng chữ
        - Row 37: Số chứng từ kèm theo
    - Signatures (row 39+ labels, signatures below):
        - B: NGƯỜI LẬP PHIẾU
        - D: NGƯỜI NHẬN HÀNG
        - F: THỦ KHO
        - H: GIÁM ĐỐC
    """
    template_path = _get_template_path(data.voucher_type)
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    # Load template
    wb = load_workbook(template_path)
    ws = wb.active
    
    # ============================================
    # FILL HEADER
    # ============================================
    # Số phiếu (ô F7, sau label "Số phiếu:" ở D7)
    ws['F7'] = data.voucher_no
    
    # Ngày (ô D2 có formula =TODAY(), ghi đè bằng ngày thực)
    ws['D2'] = _format_date(data.voucher_date)
    
    # Tên người nhận / Nhà cung cấp (ô C9)
    ws['C9'] = data.partner_name
    
    # Hóa đơn số (ô E10, sau label "Theo hóa đơn số:")
    if data.invoice_no:
        ws['E10'] = data.invoice_no
    
    # Kho (ô C11, sau label "Nhận tại kho:")
    ws['C11'] = data.warehouse_code
    
    # Địa điểm kho (ô F11)
    if data.warehouse_location:
        ws['F11'] = data.warehouse_location
    
    # ============================================
    # FILL ITEMS TABLE (rows 16-33)
    # ============================================
    start_row = 16  # Row đầu tiên của items data
    max_rows = 18   # Số dòng tối đa (16-33)
    
    total_amount = 0
    
    for i in range(max_rows):
        row = start_row + i
        
        if i < len(data.items):
            item = data.items[i]
            line_total = item.qty_actual * item.unit_price
            total_amount += line_total
            
            ws[f'A{row}'] = i + 1                    # STT
            ws[f'B{row}'] = item.name                # Tên hàng (cột B, merged với C)
            ws[f'D{row}'] = item.sku                 # Mã số
            ws[f'E{row}'] = item.unit                # Đơn vị tính
            ws[f'F{row}'] = item.qty_doc             # SL Chứng từ
            ws[f'G{row}'] = item.qty_actual          # SL Thực nhập
            ws[f'H{row}'] = item.unit_price          # Đơn giá
            # Cột I có formula tính thành tiền, không ghi đè
        else:
            # Clear unused rows (giữ nguyên format)
            ws[f'A{row}'] = None
            ws[f'B{row}'] = None
            ws[f'D{row}'] = None
            ws[f'E{row}'] = None
            ws[f'F{row}'] = None
            ws[f'G{row}'] = None
            ws[f'H{row}'] = None
    
    # ============================================
    # FILL TAX & SUMMARY (rows 34-37)
    # ============================================
    # Row 34: Thuế (template có label "Thuế 8%" - ta update theo tax_rate thực tế)
    tax_rate = data.tax_rate or 0
    tax_amount = total_amount * tax_rate / 100
    grand_total = total_amount + tax_amount
    
    # Update label thuế (cột B34)
    ws['B34'] = f"Thuế {tax_rate:.0f}%" if tax_rate > 0 else "Thuế 0%"
    # Có thể ghi tiền thuế vào cột I34 nếu template cho phép
    ws['I34'] = tax_amount if tax_rate > 0 else None
    
    # Row 35: Cộng (subtotal) - template có formula SUM, ta để nguyên
    # Nhưng cần ghi tổng tiền sau thuế vào đâu đó
    
    # Số tiền bằng chữ (row 36) - dùng grand_total (tổng sau thuế)
    total_in_words = amount_to_vietnamese_words(grand_total)
    ws['C36'] = total_in_words
    
    # Số chứng từ kèm theo (row 37)
    if data.attachments:
        ws['C37'] = data.attachments
    
    # Chữ ký (row 43-45, below labels at row 39)
    if data.prepared_by:
        ws['B43'] = data.prepared_by
    if data.receiver:
        ws['D43'] = data.receiver
    if data.storekeeper:
        ws['F43'] = data.storekeeper
    if data.director:
        ws['H43'] = data.director
    
    # ============================================
    # SAVE TO BUFFER
    # ============================================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = _generate_filename(data, 'xlsx')
    
    return buffer, filename


# ============================================
# PDF EXPORT
# ============================================

def export_to_pdf(data: VoucherExportRequest) -> tuple[io.BytesIO, str]:
    """
    Export voucher data to PDF
    
    Sử dụng reportlab để tạo PDF với layout tương tự phiếu
    
    Args:
        data: VoucherExportRequest with all voucher data
    
    Returns:
        Tuple of (BytesIO buffer, filename)
    """
    buffer = io.BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    # Custom styles (ASCII-safe for compatibility)
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        alignment=1,  # Center
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=5
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        spaceAfter=3
    )
    
    # Build document content
    elements = []
    
    # ============================================
    # HEADER
    # ============================================
    elements.append(Paragraph("CONG TY TNHH QUAN LY KHO N3T", header_style))
    elements.append(Paragraph("Dia chi: Duong Dai hoc, P. Linh Trung, TP. Thu Duc", normal_style))
    elements.append(Spacer(1, 10))
    
    # Title
    title = "PHIEU NHAP KHO" if data.voucher_type == 'PN' else "PHIEU XUAT KHO"
    elements.append(Paragraph(f"<b>{title}</b>", title_style))
    elements.append(Spacer(1, 5))
    
    # ============================================
    # VOUCHER INFO
    # ============================================
    info_data = [
        ['So phieu:', data.voucher_no, 'Ngay:', _format_date(data.voucher_date)],
        ['Doi tac:', data.partner_name, 'Kho:', data.warehouse_code],
    ]
    
    if data.invoice_no:
        info_data.append(['So hoa don:', data.invoice_no, 'Ngay HD:', _format_date(data.invoice_date) if data.invoice_date else ''])
    
    info_table = Table(info_data, colWidths=[70, 150, 60, 150])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))
    
    # ============================================
    # ITEMS TABLE
    # ============================================
    # Header
    table_header = ['STT', 'Ten hang hoa', 'Ma hang', 'DVT', 'SL CT', 'SL TT', 'Don gia', 'Thanh tien']
    
    # Data rows
    table_data = [table_header]
    total_amount = 0
    
    for i, item in enumerate(data.items):
        line_total = item.qty_actual * item.unit_price
        total_amount += line_total
        
        table_data.append([
            str(i + 1),
            item.name[:30] + '...' if len(item.name) > 30 else item.name,  # Truncate long names
            item.sku,
            item.unit,
            _format_number(item.qty_doc),
            _format_number(item.qty_actual),
            _format_number(item.unit_price),
            _format_number(line_total)
        ])
    
    # Create table
    col_widths = [25, 120, 50, 35, 40, 40, 60, 70]
    items_table = Table(table_data, colWidths=col_widths)
    items_table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Data alignment
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # STT
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),  # Numbers
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 10))
    
    # ============================================
    # SUMMARY WITH TAX
    # ============================================
    tax_rate = data.tax_rate or 0
    tax_amount = total_amount * tax_rate / 100
    grand_total = total_amount + tax_amount
    
    # Subtotal row
    summary_data = [
        ['', '', '', '', '', '', 'Cong:', _format_number(total_amount)],
    ]
    
    # Tax row (only if tax_rate > 0)
    if tax_rate > 0:
        summary_data.append(['', '', '', '', '', '', f'Thue ({tax_rate:.0f}%):', _format_number(tax_amount)])
        summary_data.append(['', '', '', '', '', '', 'TONG CONG:', _format_number(grand_total)])
    else:
        summary_data.append(['', '', '', '', '', '', 'TONG CONG:', _format_number(total_amount)])
    
    summary_table = Table(summary_data, colWidths=col_widths)
    summary_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),  # Bold last row (TONG CONG)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 5))
    
    # Total in words (use grand_total)
    total_in_words = amount_to_vietnamese_words(grand_total)
    elements.append(Paragraph(f"<i>Bang chu: {total_in_words}</i>", normal_style))
    elements.append(Spacer(1, 20))
    
    # ============================================
    # SIGNATURE
    # ============================================
    sig_data = [
        ['Nguoi lap phieu', 'Nguoi giao/nhan', 'Thu kho', 'Giam doc'],
        ['(Ky, ghi ro ho ten)', '(Ky, ghi ro ho ten)', '(Ky, ghi ro ho ten)', '(Ky, ghi ro ho ten)'],
        ['', '', '', ''],
        ['', '', '', ''],
        [data.prepared_by or '', data.receiver or '', data.storekeeper or '', data.director or ''],
    ]
    sig_table = Table(sig_data, colWidths=[110, 110, 110, 110])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, 1), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Oblique'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(sig_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = _generate_filename(data, 'pdf')
    
    return buffer, filename
